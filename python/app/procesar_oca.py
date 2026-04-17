#!/usr/bin/env python3
"""
PROCESADOR OCA - Vinculacion de Planillas a Distribuidores
Logistica Argentina SRL

USO:
  python procesar_oca.py [carpeta_con_pdfs]

Si no se indica carpeta, usa la carpeta donde esta el script.
Busca automaticamente:
  - El PDF principal de sucursal (ej: GUILLON.pdf)
  - Los PDFs individuales de cada distribuidor (ej: GUILLON Barrios Juan.pdf)

Genera: [SUCURSAL]_Vinculacion_Planillas.xlsx en la misma carpeta.

REQUISITOS (instalar una sola vez):
  pip install pdfplumber openpyxl
"""

import sys
import os
import re
from collections import defaultdict
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("ERROR: Falta instalar pdfplumber. Ejecuta:")
    print("  pip install pdfplumber")
    input("Presiona Enter para salir...")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Falta instalar openpyxl. Ejecuta:")
    print("  pip install openpyxl")
    input("Presiona Enter para salir...")
    sys.exit(1)


# ============================================================
# UTILIDADES
# ============================================================
def parse_ar_number(s):
    """Parsea numero en formato argentino: 2.165,800 -> 2165.8"""
    s = s.replace('.', '').replace(',', '.')
    return float(s)


# ============================================================
# PARSEO DE PDFs
# ============================================================
# Codigos de contrato OCA conocidos (14 totales al 17-abr-2026)
# BUGFIX 19: se agregan 197 (ROS TDC/CLEARING) y 200 (GRAL KM UTILITARIO)
CODIGOS_CONTRATO_CONOCIDOS = {
    '170': 'PAQ. ENTREGADO',
    '171': 'PAQ. MOVISTAR',
    '152': 'GRAL PICKUP',
    '181': 'PQN COMP ALTA',
    '183': 'PQN SERV MOVISTAR',
    '186': 'PQN PICKUP PRINC',
    '187': 'PQN PICKUP ADIC',
    '190': 'ADG COMPL ALTA',
    '192': 'ADG SERV MOVISTAR',
    '195': 'ADG PICKUP PRINC',
    '197': 'ROS TDC/CLEARING',           # NUEVO post-v5
    '198': 'GRAL PAQ. INTERIOR',
    '199': 'GRAL HORAS UTILITARIO',
    '200': 'GRAL KM UTILITARIO',         # NUEVO post-v5
}


def parse_main_pdf(filepath):
    """Parsea el PDF principal de sucursal con todas las planillas.

    BUGFIX 19 (Bug 2): regex GENERICO que captura cualquier codigo de contrato.
    Patron: FECHA HORA PLANILLA COD - TEXTO_CONTRATO $PRECIO QTY $TOTAL
    Los codigos desconocidos se capturan igual y se reportan como warnings.
    """
    planillas = []
    codigos_encontrados = set()

    # Regex generico: captura cualquier codigo numerico
    generic_regex = re.compile(
        r'(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})\s+([\w-]+)\s+(\d+)\s*-\s*(.+?)\s+'
        r'\$?([\d.,]+)\s+([\d.,]+)\s+\$([\d.,]+)\s*$'
    )

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                m = generic_regex.match(line.strip())
                if not m:
                    # Fallback: PICKUP malformadas (texto pegado sin espacios)
                    pickup_match = re.match(
                        r'(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})\s+([\w-]+)\s+(\d+)\s*-\s*.*?PICKUP.*?\s+'
                        r'([\d.,]+)\s+\$([\d.,]+)',
                        line
                    )
                    if pickup_match:
                        try:
                            qty_raw = parse_ar_number(pickup_match.group(5))
                            total = parse_ar_number(pickup_match.group(6))
                        except (ValueError, ZeroDivisionError):
                            continue
                        desc_text = 'PICKUP PRINC' if 'PRINC' in line else 'PICKUP ADIC' if 'ADIC' in line else 'PICKUP'
                        codigo = pickup_match.group(4)
                        codigos_encontrados.add(codigo)
                        planillas.append({
                            'fecha': pickup_match.group(1),
                            'hora': pickup_match.group(2),
                            'nro_planilla': pickup_match.group(3),
                            'cod_contrato': codigo,
                            'desc': desc_text,
                            'precio': total / qty_raw if qty_raw > 0 else 0,
                            'qty': qty_raw,
                            'total': total,
                        })
                    continue

                try:
                    qty_raw = parse_ar_number(m.group(7))
                    precio = parse_ar_number(m.group(6))
                    total = parse_ar_number(m.group(8))
                except ValueError:
                    continue

                codigo = m.group(4)
                desc_cruda = m.group(5).strip()
                codigos_encontrados.add(codigo)

                planillas.append({
                    'fecha': m.group(1),
                    'hora': m.group(2),
                    'nro_planilla': m.group(3),
                    'cod_contrato': codigo,
                    'desc': desc_cruda,
                    'precio': precio,
                    'qty': qty_raw,
                    'total': total,
                })

    # Guardar metadata de codigos para warnings
    parse_main_pdf._ultimos_codigos = codigos_encontrados
    return planillas


def parse_distributor_pdf(filepath):
    """Parsea un PDF individual de distribuidor y extrae totales diarios."""
    with pdfplumber.open(filepath) as pdf:
        full_text = ""
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    if "Total Unidades" in full_text:
        section = full_text.split("Total Unidades")[1]
    else:
        lines = full_text.split('\n')
        section_lines = []
        found_data = False
        for line in lines:
            if re.match(r'\d{2}/\d{2}/\d{4}\s*-\s*', line):
                found_data = True
                section_lines.append(line)
            elif found_data and re.match(r'\$[\d.,]+$', line.strip()):
                break
            elif found_data and line.strip() in ('', 'Total Sucursal', 'OEP', 'Planillas Manuales', 'Recorridos Fijos'):
                if line.strip() in ('Total Sucursal', 'OEP', 'Planillas Manuales', 'Recorridos Fijos'):
                    break
        section = '\n'.join(section_lines) if section_lines else full_text

    daily = {}
    for line in section.split('\n'):
        m = re.match(r'(\d{2}/\d{2}/\d{4})\s*-\s*(.+?)\s+([\d.,]+)\s+\$([\d.,]+)', line)
        if m:
            fecha = m.group(1)
            qty = parse_ar_number(m.group(3))
            amount = parse_ar_number(m.group(4))
            if fecha not in daily:
                daily[fecha] = {'qty': 0, 'amount': 0}
            daily[fecha]['qty'] += qty
            daily[fecha]['amount'] += amount
    return daily


# BUGFIX 19 (Bug 1): keywords que cierran una seccion de distribuidor.
# Al encontrar cualquiera de estas, ignorar las siguientes lineas hasta el proximo distribuidor.
# Esto evita duplicacion 2x en PDFs combinados (ROSARIO, PQO nuevo formato).
SECTION_CLOSERS = [
    'Total Unidades',
    'Total Sucursal',
    'Planillas Manuales',
    'Recorridos Fijos',
]


def parse_combined_distributor_pdf(filepath):
    """
    Parsea un PDF UNICO que contiene el desglose de TODOS los distribuidores.

    BUGFIX 19 (Bug 1): solo procesa la PRIMERA seccion de datos de cada distribuidor.
    Las secciones 'OEP Cantidad Importe' y 'Total Unidades Cantidad Importe' son
    duplicadas. Al detectar un SECTION_CLOSER, ignora todo hasta el proximo header
    de distribuidor.

    Retorna: dict {nombre_distribuidor: {fecha: {qty, amount}}}
    """
    with pdfplumber.open(filepath) as pdf:
        full_text = ""
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    lines = full_text.split('\n')
    distributors = {}
    current_dist = None
    dist_section_closed = False  # BUGFIX 19: flag de cierre de seccion

    skip_patterns = [
        'Detalle de Unidades de Recorrido',
        'Fecha de impresi',
        'Periodo:',
        'OCA',
        'LOGISTICA ARGENTINA',
        'Cantidad Importe',
        'Pagina ',
        'OEP',
    ]

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        if re.match(r'^\$[\d.,]+$', stripped):
            continue

        # Detectar header de distribuidor (texto en mayusculas)
        if not any(pat in stripped for pat in skip_patterns):
            if re.match(r'^[A-Z\u00C0-\u00DC][A-Z\u00C0-\u00DC\s.]+$', stripped) and len(stripped) > 3:
                # Nuevo distribuidor: resetear flag de cierre
                current_dist = stripped
                dist_section_closed = False
                continue

        # Detectar cierre de seccion (BUGFIX 19)
        if any(closer in stripped for closer in SECTION_CLOSERS):
            dist_section_closed = True
            continue

        # Si la seccion del distribuidor actual esta cerrada, ignorar todo
        if dist_section_closed:
            continue

        # Parsear linea de datos
        data_match = re.match(r'(\d{2}/\d{2}/\d{4})\s*-\s*(.+?)\s+([\d.,]+)\s+\$([\d.,]+)', stripped)
        if data_match and current_dist:
            fecha = data_match.group(1)
            try:
                qty = parse_ar_number(data_match.group(3))
                amount = parse_ar_number(data_match.group(4))
            except ValueError:
                continue
            if amount == 0:
                continue
            if current_dist not in distributors:
                distributors[current_dist] = {}
            if fecha not in distributors[current_dist]:
                distributors[current_dist][fecha] = {'qty': 0, 'amount': 0}
            distributors[current_dist][fecha]['qty'] += qty
            distributors[current_dist][fecha]['amount'] += amount

    return distributors


# ============================================================
# ALGORITMO DE MATCHING (Subset-Sum con Backtracking)
# ============================================================
def find_subsets_for_qty(items, target_qty, max_results=500):
    """Encuentra todos los subconjuntos de items cuya qty suma target_qty."""
    results = []
    n = len(items)
    QTY_TOLERANCE = 0.01

    def search(pos, remaining, chosen):
        if abs(remaining) < QTY_TOLERANCE:
            results.append(chosen[:])
            return len(results) >= max_results
        if remaining < -QTY_TOLERANCE or pos >= n:
            return False
        max_possible = sum(items[j]['qty'] for j in range(pos, n))
        if max_possible < remaining - QTY_TOLERANCE:
            return False
        chosen.append(pos)
        if search(pos + 1, remaining - items[pos]['qty'], chosen):
            return True
        chosen.pop()
        if search(pos + 1, remaining, chosen):
            return True
        return False

    search(0, target_qty, [])
    return results


def find_best_partition(day_items, targets):
    """
    Particiona las planillas del dia asignandolas a distribuidores.
    """
    n = len(day_items)
    best_result = None
    best_score = float('inf')

    from itertools import permutations

    if len(targets) <= 5:
        orderings = list(permutations(targets))
    else:
        orderings = [
            sorted(targets, key=lambda x: x[1]),
            sorted(targets, key=lambda x: -x[1]),
            sorted(targets, key=lambda x: x[2]),
            sorted(targets, key=lambda x: -x[2]),
        ]

    for sorted_targets in orderings:
        if best_score < 0.01:
            break

        def backtrack(tidx, used_set, current_assignment, current_score):
            nonlocal best_result, best_score
            if tidx == len(sorted_targets):
                if current_score < best_score:
                    best_score = current_score
                    best_result = dict(current_assignment)
                return
            name, target_qty, target_amount = sorted_targets[tidx]
            available = [(i, day_items[i]) for i in range(n) if i not in used_set]
            if not available:
                if target_qty == 0:
                    current_assignment[name] = []
                    backtrack(tidx + 1, used_set, current_assignment, current_score)
                    del current_assignment[name]
                return
            avail_indices = [i for i, _ in available]
            avail_items = [item for _, item in available]
            subsets = find_subsets_for_qty(avail_items, target_qty, max_results=500)
            if not subsets:
                return

            scored = []
            for subset in subsets:
                real_indices = [avail_indices[j] for j in subset]
                subset_amount = sum(day_items[i]['total'] for i in real_indices)
                amt_diff = abs(subset_amount - target_amount)
                scored.append((amt_diff, real_indices))
            scored.sort(key=lambda x: x[0])

            exact_matches = [s for s in scored if s[0] < 1.0]
            candidates = exact_matches if exact_matches else scored[:50]

            for amt_diff, real_indices in candidates:
                if current_score + amt_diff >= best_score:
                    break
                new_used = used_set | set(real_indices)
                current_assignment[name] = real_indices
                backtrack(tidx + 1, new_used, current_assignment, current_score + amt_diff)
                del current_assignment[name]

        backtrack(0, set(), {}, 0)

    return best_result, best_score


# ============================================================
# GENERACION DEL EXCEL
# ============================================================
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
DATA_FONT = Font(name='Arial', size=10)
TOTAL_FONT = Font(name='Arial', bold=True, size=10)
TOTAL_FILL = PatternFill('solid', fgColor='E2EFDA')
EXACT_FILL = PatternFill('solid', fgColor='C6EFCE')
DIFF_FILL = PatternFill('solid', fgColor='FFC7CE')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

DIST_COLORS_LIST = [
    PatternFill('solid', fgColor='BDD7EE'),
    PatternFill('solid', fgColor='FCE4D6'),
    PatternFill('solid', fgColor='E2EFDA'),
    PatternFill('solid', fgColor='D9E2F3'),
    PatternFill('solid', fgColor='FFF2CC'),
    PatternFill('solid', fgColor='F2DCDB'),
    PatternFill('solid', fgColor='D5E8D4'),
    PatternFill('solid', fgColor='DAE8FC'),
]
TAB_COLORS = ['4472C4', 'ED7D31', '70AD47', '7030A0', 'FFC000', 'C00000', '00B0F0', '808080']
SIN_ASIGNAR_FILL = PatternFill('solid', fgColor='FF9999')


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, font=None):
    cell.font = font or DATA_FONT
    cell.border = THIN_BORDER


def build_excel(planillas, by_date, distributors, dist_names, all_matches, planilla_distributor, output_path, sucursal_name):
    """Genera el Excel completo con todas las hojas."""
    wb = Workbook()
    all_dates = sorted(by_date.keys())

    dist_colors = {}
    for i, name in enumerate(dist_names):
        dist_colors[name] = DIST_COLORS_LIST[i % len(DIST_COLORS_LIST)]
    dist_colors['SIN ASIGNAR'] = SIN_ASIGNAR_FILL

    # ---- SHEET 1: Planillas + Distribuidor ----
    ws1 = wb.active
    ws1.title = "Planillas + Distribuidor"
    ws1.sheet_properties.tabColor = "1F4E79"

    headers1 = ['Fecha', 'Hora', 'Nro Planilla', 'Cod Contrato', 'Descripcion',
                'Precio Unitario', 'Cantidad', 'Importe Total', 'DISTRIBUIDOR ASIGNADO', 'MATCH']
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(headers1))

    row = 2
    for p in planillas:
        key = (p['fecha'], p['nro_planilla'], p['cod_contrato'])
        dist_name = planilla_distributor.get(key, 'SIN ASIGNAR')
        ws1.cell(row=row, column=1, value=p['fecha'])
        ws1.cell(row=row, column=2, value=p['hora'])
        ws1.cell(row=row, column=3, value=int(p['nro_planilla']) if p['nro_planilla'].isdigit() else p['nro_planilla'])
        ws1.cell(row=row, column=4, value=p['cod_contrato'])
        ws1.cell(row=row, column=5, value=p['desc'])
        ws1.cell(row=row, column=6, value=p['precio'])
        ws1.cell(row=row, column=7, value=p['qty'])
        ws1.cell(row=row, column=8, value=p['total'])
        ws1.cell(row=row, column=9, value=dist_name)
        ws1.cell(row=row, column=10, value='EXACTO' if dist_name != 'SIN ASIGNAR' else 'PENDIENTE')
        for col in range(1, 11):
            cell = ws1.cell(row=row, column=col)
            style_data_cell(cell)
            if col == 9:
                cell.fill = dist_colors.get(dist_name, PatternFill())
                cell.font = Font(name='Arial', bold=True, size=10)
            if col in [6, 8]:
                cell.number_format = '#,##0.00'
        row += 1

    widths1 = [12, 8, 12, 12, 22, 14, 10, 16, 20, 12]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.auto_filter.ref = f"A1:J{row-1}"

    # ---- SHEET 2: Resumen Diario ----
    ws2 = wb.create_sheet("Resumen Diario")
    ws2.sheet_properties.tabColor = "548235"
    ws2.cell(row=1, column=1, value=f"RESUMEN DIARIO - {sucursal_name} - VINCULACION PLANILLAS")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(dist_names) * 3)
    ws2.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=14, color='1F4E79')

    ws2.cell(row=3, column=1, value='Fecha')
    col = 2
    for name in dist_names:
        ws2.cell(row=3, column=col, value=f'{name}\nCantidad')
        ws2.cell(row=3, column=col + 1, value=f'{name}\nImporte PDF')
        ws2.cell(row=3, column=col + 2, value=f'{name}\nImporte Calc')
        col += 3
    ws2.cell(row=3, column=col, value='Total Dia')
    ws2.cell(row=3, column=col + 1, value='Estado')
    style_header(ws2, 3, col + 1)
    ws2.row_dimensions[3].height = 35

    data_row = 4
    for date in all_dates:
        ws2.cell(row=data_row, column=1, value=date)
        style_data_cell(ws2.cell(row=data_row, column=1))
        col = 2
        day_ok = True
        for name in dist_names:
            target_qty = distributors[name].get(date, {}).get('qty', 0)
            target_amt = distributors[name].get(date, {}).get('amount', 0)
            calc_amt = 0
            if date in all_matches and name in all_matches[date]:
                day_items = by_date[date]
                for idx in all_matches[date][name]:
                    calc_amt += day_items[idx]['total']
            ws2.cell(row=data_row, column=col, value=target_qty)
            ws2.cell(row=data_row, column=col + 1, value=target_amt)
            ws2.cell(row=data_row, column=col + 2, value=calc_amt)
            for c in range(col, col + 3):
                style_data_cell(ws2.cell(row=data_row, column=c))
                ws2.cell(row=data_row, column=c).fill = dist_colors[name]
            ws2.cell(row=data_row, column=col + 1).number_format = '#,##0.00'
            ws2.cell(row=data_row, column=col + 2).number_format = '#,##0.00'
            diff = abs(target_amt - calc_amt)
            if diff > 1 and target_amt > 0:
                day_ok = False
                ws2.cell(row=data_row, column=col + 2).fill = DIFF_FILL
            col += 3
        total_day = sum(p['total'] for p in by_date[date])
        ws2.cell(row=data_row, column=col, value=total_day)
        ws2.cell(row=data_row, column=col).number_format = '#,##0.00'
        style_data_cell(ws2.cell(row=data_row, column=col))
        ws2.cell(row=data_row, column=col + 1, value='OK' if day_ok else 'REVISAR')
        style_data_cell(ws2.cell(row=data_row, column=col + 1))
        ws2.cell(row=data_row, column=col + 1).fill = EXACT_FILL if day_ok else DIFF_FILL
        ws2.cell(row=data_row, column=col + 1).font = Font(name='Arial', bold=True, size=10)
        data_row += 1

    # Totals
    ws2.cell(row=data_row, column=1, value='TOTAL')
    ws2.cell(row=data_row, column=1).font = TOTAL_FONT
    ws2.cell(row=data_row, column=1).fill = TOTAL_FILL
    col = 2
    for name in dist_names:
        total_qty = sum(distributors[name].get(d, {}).get('qty', 0) for d in all_dates)
        total_amt = sum(distributors[name].get(d, {}).get('amount', 0) for d in all_dates)
        ws2.cell(row=data_row, column=col, value=total_qty)
        ws2.cell(row=data_row, column=col + 1, value=total_amt)
        ws2.cell(row=data_row, column=col + 1).number_format = '#,##0.00'
        for c in range(col, col + 3):
            ws2.cell(row=data_row, column=c).font = TOTAL_FONT
            ws2.cell(row=data_row, column=c).fill = TOTAL_FILL
            ws2.cell(row=data_row, column=c).border = THIN_BORDER
        col += 3
    ws2.column_dimensions['A'].width = 12
    for i in range(2, col + 2):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    # ---- SHEETS POR DISTRIBUIDOR ----
    for di, dist_name in enumerate(dist_names):
        short_name = dist_name.split()[0] if len(dist_name.split()) > 1 else dist_name[:10]
        ws = wb.create_sheet(short_name)
        ws.sheet_properties.tabColor = TAB_COLORS[di % len(TAB_COLORS)]

        ws.cell(row=1, column=1, value=f"DETALLE PLANILLAS - {dist_name.upper()}")
        ws.merge_cells('A1:H1')
        ws.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=14, color='1F4E79')

        headers = ['Fecha', 'Hora', 'Nro Planilla', 'Cod Contrato', 'Descripcion',
                   'Precio Unit.', 'Cantidad', 'Importe']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))

        row = 4
        for date in all_dates:
            if date not in all_matches or dist_name not in all_matches[date]:
                continue
            day_items = by_date[date]
            indices = all_matches[date][dist_name]
            matched = sorted([day_items[i] for i in indices], key=lambda x: str(x['nro_planilla']))
            for p in matched:
                ws.cell(row=row, column=1, value=p['fecha'])
                ws.cell(row=row, column=2, value=p['hora'])
                ws.cell(row=row, column=3, value=int(p['nro_planilla']) if p['nro_planilla'].isdigit() else p['nro_planilla'])
                ws.cell(row=row, column=4, value=p['cod_contrato'])
                ws.cell(row=row, column=5, value=p['desc'])
                ws.cell(row=row, column=6, value=p['precio'])
                ws.cell(row=row, column=7, value=p['qty'])
                ws.cell(row=row, column=8, value=p['total'])
                for c in range(1, 9):
                    style_data_cell(ws.cell(row=row, column=c))
                    if c in [6, 8]:
                        ws.cell(row=row, column=c).number_format = '#,##0.00'
                row += 1

            day_qty = sum(day_items[i]['qty'] for i in indices)
            day_amt = sum(day_items[i]['total'] for i in indices)
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=5, value='SUBTOTAL DIA')
            ws.cell(row=row, column=7, value=day_qty)
            ws.cell(row=row, column=8, value=day_amt)
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            for c in range(1, 9):
                ws.cell(row=row, column=c).font = TOTAL_FONT
                ws.cell(row=row, column=c).fill = dist_colors[dist_name]
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

            target_qty = distributors[dist_name].get(date, {}).get('qty', 0)
            target_amt = distributors[dist_name].get(date, {}).get('amount', 0)
            ws.cell(row=row, column=5, value='vs PDF Distribuidor')
            ws.cell(row=row, column=7, value=target_qty)
            ws.cell(row=row, column=8, value=target_amt)
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            for c in [5, 7, 8]:
                ws.cell(row=row, column=c).font = Font(name='Arial', italic=True, size=9, color='666666')
                ws.cell(row=row, column=c).border = THIN_BORDER
            diff = abs(day_amt - target_amt)
            ws.cell(row=row, column=8).fill = EXACT_FILL if diff < 1 else DIFF_FILL
            row += 1

        row += 1
        calc_qty = 0
        calc_amt = 0
        for date in all_dates:
            if date in all_matches and dist_name in all_matches[date]:
                day_items = by_date[date]
                for idx in all_matches[date][dist_name]:
                    calc_qty += day_items[idx]['qty']
                    calc_amt += day_items[idx]['total']
        total_qty_pdf = sum(distributors[dist_name].get(d, {}).get('qty', 0) for d in all_dates)
        total_amt_pdf = sum(distributors[dist_name].get(d, {}).get('amount', 0) for d in all_dates)

        for label, qty_val, amt_val, is_diff_row in [
            ('TOTAL CALCULADO', calc_qty, calc_amt, False),
            ('TOTAL PDF', total_qty_pdf, total_amt_pdf, False),
            ('DIFERENCIA', calc_qty - total_qty_pdf, calc_amt - total_amt_pdf, True),
        ]:
            ws.cell(row=row, column=5, value=label)
            ws.cell(row=row, column=7, value=qty_val)
            ws.cell(row=row, column=8, value=amt_val)
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            fill = (EXACT_FILL if abs(amt_val) < 10 else DIFF_FILL) if is_diff_row else TOTAL_FILL
            for c in range(1, 9):
                ws.cell(row=row, column=c).font = Font(name='Arial', bold=True, size=11)
                ws.cell(row=row, column=c).fill = fill
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        widths = [12, 8, 12, 12, 20, 14, 10, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- SHEET: Validacion Cruzada ----
    ws_val = wb.create_sheet("Validacion Cruzada")
    ws_val.sheet_properties.tabColor = "FF0000"
    ws_val.cell(row=1, column=1, value="VALIDACION CRUZADA - CONTROL DE INTEGRIDAD")
    ws_val.merge_cells('A1:H1')
    ws_val.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=14, color='C00000')

    headers_val = ['Distribuidor', 'Qty PDF', 'Qty Calculado', 'Diff Qty',
                   'Importe PDF', 'Importe Calculado', 'Diff Importe', 'Estado']
    for col, h in enumerate(headers_val, 1):
        ws_val.cell(row=3, column=col, value=h)
    style_header(ws_val, 3, len(headers_val))

    row = 4
    gt_pdf_qty = gt_calc_qty = gt_pdf_amt = gt_calc_amt = 0
    for name in dist_names:
        total_qty_pdf = sum(distributors[name].get(d, {}).get('qty', 0) for d in all_dates)
        total_amt_pdf = sum(distributors[name].get(d, {}).get('amount', 0) for d in all_dates)
        calc_qty = calc_amt = 0
        for date in all_dates:
            if date in all_matches and name in all_matches[date]:
                day_items = by_date[date]
                for idx in all_matches[date][name]:
                    calc_qty += day_items[idx]['qty']
                    calc_amt += day_items[idx]['total']
        ws_val.cell(row=row, column=1, value=name)
        ws_val.cell(row=row, column=2, value=total_qty_pdf)
        ws_val.cell(row=row, column=3, value=calc_qty)
        ws_val.cell(row=row, column=4, value=calc_qty - total_qty_pdf)
        ws_val.cell(row=row, column=5, value=total_amt_pdf)
        ws_val.cell(row=row, column=6, value=calc_amt)
        ws_val.cell(row=row, column=7, value=calc_amt - total_amt_pdf)
        diff = abs(calc_amt - total_amt_pdf)
        ws_val.cell(row=row, column=8, value='OK' if diff < 10 else f'DIFF ${diff:,.2f}')
        for c in range(1, 9):
            style_data_cell(ws_val.cell(row=row, column=c))
            if c in [5, 6, 7]:
                ws_val.cell(row=row, column=c).number_format = '#,##0.00'
        ws_val.cell(row=row, column=1).fill = dist_colors[name]
        ws_val.cell(row=row, column=8).fill = EXACT_FILL if diff < 10 else DIFF_FILL
        ws_val.cell(row=row, column=8).font = Font(name='Arial', bold=True, size=10)
        gt_pdf_qty += total_qty_pdf
        gt_calc_qty += calc_qty
        gt_pdf_amt += total_amt_pdf
        gt_calc_amt += calc_amt
        row += 1

    row += 1
    ws_val.cell(row=row, column=1, value='TOTAL SUCURSAL')
    ws_val.cell(row=row, column=2, value=gt_pdf_qty)
    ws_val.cell(row=row, column=3, value=gt_calc_qty)
    ws_val.cell(row=row, column=4, value=gt_calc_qty - gt_pdf_qty)
    ws_val.cell(row=row, column=5, value=gt_pdf_amt)
    ws_val.cell(row=row, column=6, value=gt_calc_amt)
    ws_val.cell(row=row, column=7, value=gt_calc_amt - gt_pdf_amt)
    grand_diff = abs(gt_calc_amt - gt_pdf_amt)
    ws_val.cell(row=row, column=8, value='OK' if grand_diff < 50 else 'REVISAR')
    for c in range(1, 9):
        ws_val.cell(row=row, column=c).font = TOTAL_FONT
        ws_val.cell(row=row, column=c).fill = TOTAL_FILL
        ws_val.cell(row=row, column=c).border = THIN_BORDER
        if c in [5, 6, 7]:
            ws_val.cell(row=row, column=c).number_format = '#,##0.00'

    row += 2
    main_total_qty = sum(p['qty'] for p in planillas)
    main_total_amt = sum(p['total'] for p in planillas)
    ws_val.cell(row=row, column=1, value='Total PDF Principal')
    ws_val.cell(row=row, column=2, value=main_total_qty)
    ws_val.cell(row=row, column=5, value=main_total_amt)
    ws_val.cell(row=row, column=5).number_format = '#,##0.00'
    for c in [1, 2, 5]:
        ws_val.cell(row=row, column=c).font = Font(name='Arial', bold=True, size=10, color='1F4E79')
        ws_val.cell(row=row, column=c).border = THIN_BORDER

    widths_val = [20, 12, 14, 10, 18, 18, 18, 16]
    for i, w in enumerate(widths_val, 1):
        ws_val.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    return wb


# ============================================================
# DETECCION AUTOMATICA DE ARCHIVOS
# ============================================================
def detect_files(folder):
    """Detecta automaticamente el PDF principal y los de distribuidores."""
    pdfs = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]

    if not pdfs:
        print(f"ERROR: No se encontraron archivos PDF en {folder}")
        return None, None, None

    main_pdf = None
    dist_pdfs = {}
    combined_pdf = None

    prefixes = defaultdict(list)
    for pdf in pdfs:
        base = pdf.replace('.pdf', '').replace('.PDF', '')
        parts = base.split(' ', 1)
        prefixes[parts[0]].append(pdf)

    sucursal = max(prefixes.keys(), key=lambda k: len(prefixes[k]))

    for pdf in prefixes[sucursal]:
        base = pdf.replace('.pdf', '').replace('.PDF', '')
        if ' ' not in base:
            main_pdf = pdf
        elif 'desglose' in base.lower():
            combined_pdf = pdf
        else:
            dist_name = base.replace(sucursal + ' ', '')
            dist_pdfs[dist_name] = pdf

    if not main_pdf:
        print(f"ERROR: No se encontro PDF principal para sucursal {sucursal}")
        return None, None, None

    if not dist_pdfs and combined_pdf:
        print(f"  Modo: PDF de desglose unico ({combined_pdf})")
        return main_pdf, combined_pdf, sucursal
    elif dist_pdfs:
        print(f"  Modo: PDFs individuales por distribuidor")
        return main_pdf, dist_pdfs, sucursal
    else:
        print(f"ERROR: No se encontraron PDFs de distribuidores para {sucursal}")
        return None, None, None


# ============================================================
# MAIN (standalone)
# ============================================================
def main():
    if len(sys.argv) > 1:
        folder = sys.argv[1]
    else:
        folder = os.path.dirname(os.path.abspath(__file__))

    if not os.path.isdir(folder):
        print(f"ERROR: La carpeta '{folder}' no existe")
        input("Presiona Enter para salir...")
        return

    print(f"=" * 60)
    print(f"  PROCESADOR OCA - Logistica Argentina SRL")
    print(f"=" * 60)
    print(f"\nCarpeta: {folder}")

    main_pdf, dist_pdfs_or_combined, sucursal = detect_files(folder)
    if not main_pdf:
        input("Presiona Enter para salir...")
        return

    print(f"\nSucursal detectada: {sucursal}")
    print(f"PDF principal: {main_pdf}")

    is_combined_mode = isinstance(dist_pdfs_or_combined, str)

    if is_combined_mode:
        print(f"PDF de desglose: {dist_pdfs_or_combined}")
    else:
        print(f"Distribuidores encontrados: {len(dist_pdfs_or_combined)}")
        for name, pdf in sorted(dist_pdfs_or_combined.items()):
            print(f"  - {name} ({pdf})")

    print(f"\nProcesando PDF principal...")
    planillas = parse_main_pdf(os.path.join(folder, main_pdf))
    print(f"  {len(planillas)} planillas encontradas")

    by_date = defaultdict(list)
    for p in planillas:
        by_date[p['fecha']].append(p)
    print(f"  {len(by_date)} dias con operaciones")

    distributors = {}
    if is_combined_mode:
        print(f"\nProcesando PDF de desglose unico...")
        combined_path = os.path.join(folder, dist_pdfs_or_combined)
        distributors = parse_combined_distributor_pdf(combined_path)
        dist_names = sorted(distributors.keys())
        print(f"  Distribuidores detectados: {len(dist_names)}")
        for name in dist_names:
            total_qty = sum(d['qty'] for d in distributors[name].values())
            total_amt = sum(d['amount'] for d in distributors[name].values())
            print(f"  {name}: {len(distributors[name])} dias, {total_qty:.0f} paquetes, ${total_amt:,.2f}")
    else:
        print(f"\nProcesando PDFs de distribuidores...")
        dist_names = sorted(dist_pdfs_or_combined.keys())
        for name in dist_names:
            filepath = os.path.join(folder, dist_pdfs_or_combined[name])
            distributors[name] = parse_distributor_pdf(filepath)
            total_qty = sum(d['qty'] for d in distributors[name].values())
            total_amt = sum(d['amount'] for d in distributors[name].values())
            print(f"  {name}: {len(distributors[name])} dias, {total_qty:.0f} paquetes, ${total_amt:,.2f}")

    print(f"\nVerificando totales diarios...")
    all_dates = sorted(by_date.keys())
    mismatches = 0
    for date in all_dates:
        main_qty = sum(p['qty'] for p in by_date[date])
        dist_qty = sum(d.get(date, {}).get('qty', 0) for d in distributors.values())
        if main_qty != dist_qty:
            print(f"  ADVERTENCIA: {date} - Principal={main_qty} vs Distribuidores={dist_qty}")
            mismatches += 1
    if mismatches == 0:
        print(f"  Todos los dias coinciden OK")

    print(f"\nEjecutando algoritmo de vinculacion...")
    all_matches = {}
    exact_days = 0
    for date in all_dates:
        day_items = by_date[date]
        targets = [(name, distributors[name][date]['qty'], distributors[name][date]['amount'])
                    for name in dist_names if date in distributors[name]]
        if not targets:
            continue
        result, score = find_best_partition(day_items, targets)
        if result:
            all_matches[date] = result
            if score < 1:
                exact_days += 1
            print(f"  {date}: {'EXACTO' if score < 1 else f'score={score:.2f}'}")
        else:
            print(f"  {date}: SIN SOLUCION")

    print(f"\n  Resumen: {exact_days}/{len(all_dates)} dias con match exacto")

    planilla_distributor = {}
    for date, partition in all_matches.items():
        day_items = by_date[date]
        for name, indices in partition.items():
            for idx in indices:
                p = day_items[idx]
                key = (p['fecha'], p['nro_planilla'], p['cod_contrato'])
                planilla_distributor[key] = name

    matched = len(planilla_distributor)
    total = len(planillas)
    print(f"  Planillas vinculadas: {matched}/{total} ({100*matched/total:.1f}%)")

    output_name = f"{sucursal}_Vinculacion_Planillas.xlsx"
    output_path = os.path.join(folder, output_name)
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except PermissionError:
            output_name = f"{sucursal}_Vinculacion_Planillas_NUEVO.xlsx"
            output_path = os.path.join(folder, output_name)
            print(f"  NOTA: El archivo anterior esta abierto. Guardando como: {output_name}")
    print(f"\nGenerando Excel: {output_name}")
    build_excel(planillas, by_date, distributors, dist_names, all_matches,
                planilla_distributor, output_path, sucursal)
    print(f"  Guardado en: {output_path}")

    print(f"\n{'=' * 60}")
    print(f"  PROCESO COMPLETADO")
    print(f"  Archivo: {output_name}")
    print(f"  Planillas: {matched}/{total} vinculadas")
    print(f"  Dias exactos: {exact_days}/{len(all_dates)}")
    print(f"{'=' * 60}")

    input("\nPresiona Enter para salir...")


if __name__ == '__main__':
    main()
